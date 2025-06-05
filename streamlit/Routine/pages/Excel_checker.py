import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Any

st.set_page_config(page_title="File Check", page_icon="🔍")

def get_cell_visibility_info(ws) -> Dict[str, List[int]]:
    """セルの表示/非表示情報を取得"""
    hidden_rows = [i for i in range(1, ws.max_row + 1) 
                  if ws.row_dimensions[i].hidden]
    hidden_cols = [i for i in range(1, ws.max_column + 1) 
                  if ws.column_dimensions[get_column_letter(i)].hidden]
    return {'hidden_rows': hidden_rows, 'hidden_cols': hidden_cols}

def read_file(file) -> Optional[Dict[str, Any]]:
    """ExcelファイルまたはCSVファイルを読み込む"""
    try:
        file_extension = file.name.split('.')[-1].lower()
        if file_extension in ['xlsx', 'xlsm']:
            # Excel形式の場合、DataFrameとワークブックの両方を保持
            excel_data = pd.ExcelFile(file)
            wb = load_workbook(file, data_only=True)
            return {'type': 'excel', 'file': excel_data, 'workbook': wb}
        elif file_extension == 'csv':
            return {'type': 'csv', 'file': pd.read_csv(file)}
        return None
    except Exception as e:
        st.error(f"ファイル読み込みエラー: {e}")
        return None

def get_dataframe_and_columns(file_info: Dict[str, Any], sheet_name: Optional[str] = None) -> Tuple[Optional[pd.DataFrame], Optional[Dict[str, List[int]]], Optional[Dict[str, str]]]:
    """ファイル情報からDataFrameと列の対応情報を取得"""
    try:
        if file_info['type'] == 'excel':
            # DataFrameの取得
            excel_file = file_info['file']
            if sheet_name is not None:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
            else:
                # sheet_nameが指定されていない場合は最初のシートを読み込む
                df = next(iter(pd.read_excel(excel_file, sheet_name=None).values()))
            
            ws = file_info['workbook'][sheet_name] if sheet_name else file_info['workbook'].active
            visibility_info = get_cell_visibility_info(ws)
            
            # 列の対応関係を取得
            column_mapping = {}
            for i, col_name in enumerate(df.columns):
                excel_col = get_column_letter(i + 1)
                column_mapping[str(col_name)] = excel_col
            
            return df, visibility_info, column_mapping
        else:  # CSV
            df = file_info['file']
            if isinstance(df, pd.DataFrame):  # DataFrameであることを確認
                column_mapping = {str(col): get_column_letter(i + 1) 
                                for i, col in enumerate(df.columns)}
                return df, None, column_mapping
            else:
                raise TypeError("CSVファイルの読み込みに失敗しました")
    except Exception as e:
        st.error(f"データフレーム取得エラー: {e}")
        return None, None, None

def compare_dataframes(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    sheet_name: str,
    visibility_info1: Optional[Dict[str, List[int]]] = None,
    visibility_info2: Optional[Dict[str, List[int]]] = None,
    column_mapping1: Optional[Dict[str, str]] = None,
    column_mapping2: Optional[Dict[str, str]] = None
) -> List[Dict[str, Any]]:
    """2つのDataFrameを比較して違いを見つける"""
    differences = []
    
    # 列名と行数を統一
    all_columns = list(set(df1.columns) | set(df2.columns))
    max_rows = max(len(df1), len(df2))
    
    # 不足している列を追加
    for col in all_columns:
        if col not in df1.columns:
            df1[col] = np.nan
        if col not in df2.columns:
            df2[col] = np.nan
    
    # インデックスと列を揃える
    df1 = df1.reindex(columns=all_columns)
    df2 = df2.reindex(columns=all_columns)
    
    # 各セルを比較
    for col in all_columns:
        # 列のExcel形式の参照を取得
        excel_col1 = column_mapping1.get(str(col), '') if column_mapping1 else get_column_letter(all_columns.index(col) + 1)
        
        for idx in range(max_rows):
            try:
                val1 = df1.iloc[idx][col] if idx < len(df1) else np.nan
                val2 = df2.iloc[idx][col] if idx < len(df2) else np.nan
                
                # NANの処理
                if pd.isna(val1) and pd.isna(val2):
                    continue
                
                # 値が異なる場合
                if pd.isna(val1) != pd.isna(val2) or val1 != val2:
                    row_num = idx + 1
                    cell_address = f"{excel_col1}{row_num}"
                    
                    # 非表示情報の追加
                    visibility_status = ""
                    if visibility_info1:
                        if row_num in visibility_info1['hidden_rows']:
                            visibility_status += "行: 非表示 "
                        col_idx = column_index_from_string(excel_col1)
                        if col_idx in visibility_info1['hidden_cols']:
                            visibility_status += "列: 非表示"
                    
                    differences.append({
                        'シート名': sheet_name,
                        'セル': cell_address,
                        '行': row_num,
                        '列': col,
                        'ファイル1の値': val1,
                        'ファイル2の値': val2,
                        '表示状態': visibility_status.strip() or "表示"
                    })
            except Exception as e:
                st.error(f"セル比較エラー: {e}")
    
    return differences

def compare_files(file1, file2) -> Optional[List[Dict[str, Any]]]:
    differences = []
    
    # ファイルを読み込み
    file1_info = read_file(file1)
    file2_info = read_file(file2)
    
    if not file1_info or not file2_info:
        return None

    # CSVファイルの場合は単一のDataFrameとして比較
    if file1_info['type'] == 'csv' and file2_info['type'] == 'csv':
        df1, _, column_mapping1 = get_dataframe_and_columns(file1_info)
        df2, _, column_mapping2 = get_dataframe_and_columns(file2_info)
        if df1 is not None and df2 is not None:
            differences.extend(compare_dataframes(
                df1, df2, "CSV",
                column_mapping1=column_mapping1,
                column_mapping2=column_mapping2
            ))
        
    # Excelファイルの場合はシート毎に比較
    elif file1_info['type'] == 'excel' and file2_info['type'] == 'excel':
        sheets1 = {str(sheet) for sheet in file1_info['file'].sheet_names}
        sheets2 = {str(sheet) for sheet in file2_info['file'].sheet_names}
        
        common_sheets = sheets1 & sheets2
        sheets_only_in_1 = sheets1 - sheets2
        sheets_only_in_2 = sheets2 - sheets1
        
        if sheets_only_in_1:
            st.warning(f"ファイル1のみに存在するシート: {', '.join(sheets_only_in_1)}")
        if sheets_only_in_2:
            st.warning(f"ファイル2のみに存在するシート: {', '.join(sheets_only_in_2)}")

        progress_bar = st.progress(0)
        total_sheets = len(common_sheets)
        
        for i, sheet_name in enumerate(common_sheets):
            progress = (i + 1) / total_sheets
            progress_bar.progress(progress)
            
            st.write(f"シート '{sheet_name}' を比較中...")
            
            df1, visibility_info1, column_mapping1 = get_dataframe_and_columns(file1_info, sheet_name)
            df2, visibility_info2, column_mapping2 = get_dataframe_and_columns(file2_info, sheet_name)
            
            if df1 is not None and df2 is not None:
                sheet_differences = compare_dataframes(
                    df1, df2, sheet_name, 
                    visibility_info1, visibility_info2,
                    column_mapping1, column_mapping2
                )
                differences.extend(sheet_differences)
    
    # Excel と CSV の組み合わせの場合はエラー
    else:
        st.error("ExcelファイルとCSVファイルを同時に比較することはできません。")
        return None

    return differences

# メインのUIコード
st.title('ファイル比較ツール (xlsx/xlsm/csv対応)')
st.write('xlsxファイル、xlsmファイル、またはCSVファイルを比較します')

# ファイルアップロード
file1 = st.file_uploader("1つ目のファイルをアップロード", type=['xlsx', 'xlsm', 'csv'])
file2 = st.file_uploader("2つ目のファイルをアップロード", type=['xlsx', 'xlsm', 'csv'])

if file1 and file2:
    col1, col2 = st.columns(2)
    with col1:
        st.write("ファイル1:", file1.name)
    with col2:
        st.write("ファイル2:", file2.name)

    if st.button('比較開始'):
        with st.spinner('比較中...'):
            differences = compare_files(file1, file2)
            
            if differences:
                # 結果をDataFrameに変換
                result_df = pd.DataFrame(differences)
                
                # 差分の数を表示
                st.write(f"### 見つかった差分: {len(differences)}件")
                
                # 結果を表示（ページング機能付き）
                st.dataframe(result_df)
                
                # Excelファイルとしてダウンロード可能に
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    result_df.to_excel(writer, index=False)
                
                output.seek(0)
                st.download_button(
                    label="結果をExcelファイルとしてダウンロード",
                    data=output,
                    file_name="比較結果.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.success("違いは見つかりませんでした。")
